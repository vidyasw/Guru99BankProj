# import unittest
import xlrd
from openpyxl import load_workbook
import openpyexcel


# class TC_Excel_utility(unittest.TestCase):

# @staticmethod
def get_row_count(file,sheet_name):
    workbook = openpyexcel.load_workbook(file)
    sheet = workbook[sheet_name]
    # sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.max_row


# @staticmethod
def get_column_count(file,sheet_name):
    workbook = openpyexcel.load_workbook(file)
    sheet = workbook[sheet_name]
    # sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.max_column


# @staticmethod
def read_data(file,sheet_name,row_num,column_num):
    workbook = openpyexcel.load_workbook(file)
    sheet = workbook[sheet_name]
    # sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.cell(row=row_num,column=column_num).value


# @staticmethod
def write_data(file,sheet_name,row_num,column_num,data):
    workbook = openpyexcel.load_workbook(file)
    sheet = workbook[sheet_name]
    # sheet = workbook.get_sheet_by_name(sheet_name)
    sheet.cell(row=row_num,column=column_num).value = data
    workbook.save(file)


# @staticmethod
def delete_row(file,sheet_name,row_num,no_row_delete):
    try:
        workbook = openpyexcel.load_workbook(file)
        sheet = workbook[sheet_name]
        # sheet = workbook.get_sheet_by_name(sheet_name)
        sheet.delete_rows(row_num,no_row_delete)
        workbook.save(file)
    except Exception as e:
        print("In Delete_row method:",type(e).__name__)


'''this will return row and column index for to_serach string'''
# @staticmethod
def find_row_col_as_per_serach_string(file,sheet_index,to_search_string):
    row_idx, col_idx  = 0
    book = xlrd.open_workbook(file)
    sheet = book.sheet_by_index(sheet_index)
    for row_idx in range(sheet.nrows):
        row = sheet.row(row_idx)
        for col_idx, cell in enumerate(row):
            if cell.value == to_search_string:
                print("Sheet Name :",sheet.name)
                print("Col Index:",col_idx)
                print("Row Index:",row_idx)
    return sheet.name,row_idx,col_idx

    # ws['B'] will retu
    # rn all cells on the B column until the last one
    # (similar to max_row but it's only for the B column)
    '''for cell in ws['A']:
        if cell.value is not None:
            if "58986" in str(cell.value):
                print('Found header with name: {} at row: {} and column: {}. In cell {}'.format(cell.value,cell.row,cell.column,cell))'''


# @staticmethod
def search_value_in_column(file, sheet_name,search_string, column):
    wb = load_workbook(file)
    ws = wb[sheet_name]
    for row in range(1, ws.max_row + 1):
        coordinate = "{}{}".format(column, row)
        if ws[coordinate].value == search_string:
            return column, row
    return column, None


# @staticmethod
def search_value_in_col_idx(file,sheet_name,search_string, col_idx=1):
    wb = load_workbook(file)
    ws = wb[sheet_name]
    for row in range(1, ws.max_row + 1):
        if ws[row][col_idx].value == search_string:
            return col_idx, row
    return col_idx, None


# @staticmethod
def search_value_in_row_index(file,sheet_name, search_string, row=1):
    wb = load_workbook(file)
    ws = wb[sheet_name]
    for cell in ws[row]:
        if cell.value == search_string:
            return cell.column, row
    return None, row